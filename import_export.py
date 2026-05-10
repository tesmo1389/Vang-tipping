"""
import_export.py - Excel/CSV import og eksport for VM 2026 tipping.
"""
import io
import csv
from datetime import datetime
import pytz
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from models import db, Match, Team, Group, User, UserPrediction, GroupPrediction, AdminAuditLog, now_utc
from scoring import recalculate_all_scores

OSLO_TZ = pytz.timezone("Europe/Oslo")

MATCH_COLUMNS = [
    "match_number", "phase", "round_name", "group_name",
    "kickoff_date_et", "kickoff_time_et", "kickoff_timezone_source",
    "venue", "city", "country",
    "home_team_code", "away_team_code",
    "home_score", "away_score", "is_finished", "advanced_team_code", "result_note"
]


def generate_excel_template():
    """Generate Excel template for match import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kampplan"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(MATCH_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Add example row
    example = {
        "match_number": 1,
        "phase": "group",
        "round_name": "group_stage",
        "group_name": "A",
        "kickoff_date_et": "2026-06-11",
        "kickoff_time_et": "15:00",
        "kickoff_timezone_source": "America/New_York",
        "venue": "SoFi Stadium",
        "city": "Los Angeles",
        "country": "USA",
        "home_team_code": "MEX",
        "away_team_code": "RSA",
        "home_score": "",
        "away_score": "",
        "is_finished": "FALSE",
        "advanced_team_code": "",
        "result_note": "",
    }
    for col, key in enumerate(MATCH_COLUMNS, 1):
        ws.cell(row=2, column=col, value=example.get(key, ""))

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_csv_template():
    """Generate CSV template for match import."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=MATCH_COLUMNS)
    writer.writeheader()
    writer.writerow({
        "match_number": 1,
        "phase": "group",
        "round_name": "group_stage",
        "group_name": "A",
        "kickoff_date_et": "2026-06-11",
        "kickoff_time_et": "15:00",
        "kickoff_timezone_source": "America/New_York",
        "venue": "SoFi Stadium",
        "city": "Los Angeles",
        "country": "USA",
        "home_team_code": "MEX",
        "away_team_code": "RSA",
        "home_score": "",
        "away_score": "",
        "is_finished": "FALSE",
        "advanced_team_code": "",
        "result_note": "",
    })
    return buf.getvalue().encode("utf-8")


def validate_and_preview_import(file_content, file_type="xlsx"):
    """Validate import file and return preview rows before committing."""
    rows, errors = _parse_import_file(file_content, file_type)
    if errors:
        return {"success": False, "errors": errors, "rows": []}

    preview = []
    row_errors = []
    for row in rows:
        err = _validate_row(row)
        if err:
            row_errors.append({"row": row, "error": err})
        else:
            preview.append(row)

    if row_errors:
        return {"success": False, "errors": [f"Rad {r['row'].get('match_number', '?')}: {r['error']}" for r in row_errors], "rows": []}

    return {"success": True, "errors": [], "rows": preview}


def execute_import(rows):
    """Execute validated import rows. Returns result dict."""
    new_count = 0
    updated_count = 0
    errors = []

    team_cache = {}
    group_cache = {}

    def get_team(code):
        if code not in team_cache:
            team_cache[code] = Team.query.filter_by(fifa_code=code).first()
        return team_cache[code]

    def get_group(letter):
        name = f"Gruppe {letter}"
        if name not in group_cache:
            group_cache[name] = Group.query.filter_by(name=name).first()
        return group_cache[name]

    for row in rows:
        try:
            match_number = int(row["match_number"])
            match = Match.query.filter_by(match_number=match_number).first()
            if not match:
                match = Match(match_number=match_number, phase=row.get("phase", "group"))
                db.session.add(match)
                new_count += 1
            else:
                updated_count += 1

            match.phase = row.get("phase", match.phase)
            match.round_name = row.get("round_name", match.round_name)

            group_letter = row.get("group_name", "").strip()
            if group_letter:
                grp = get_group(group_letter)
                if grp:
                    match.group_id = grp.id

            # Kickoff
            date_et = row.get("kickoff_date_et", "").strip()
            time_et = row.get("kickoff_time_et", "").strip()
            tz_src = row.get("kickoff_timezone_source", "America/New_York").strip() or "America/New_York"
            if date_et and time_et:
                from schedule_import import parse_kickoff
                from datetime import timedelta
                kickoff_utc = parse_kickoff(date_et, time_et, tz_src)
                match.kickoff_at_utc = kickoff_utc
                match.kickoff_timezone_source = tz_src
                if kickoff_utc:
                    utc_aware = pytz.utc.localize(kickoff_utc)
                    oslo_dt = utc_aware.astimezone(OSLO_TZ)
                    match.kickoff_at_oslo_cache = oslo_dt.strftime("%Y-%m-%d %H:%M")
                    if match.phase == "knockout":
                        match.lock_at_utc = kickoff_utc - timedelta(hours=1)

            match.venue = row.get("venue", match.venue)
            match.city = row.get("city", match.city)
            match.country = row.get("country", match.country)

            home_code = row.get("home_team_code", "").strip()
            away_code = row.get("away_team_code", "").strip()
            if home_code:
                t = get_team(home_code)
                if t:
                    match.home_team_id = t.id
            if away_code:
                t = get_team(away_code)
                if t:
                    match.away_team_id = t.id

            # Result
            home_score = row.get("home_score", "")
            away_score = row.get("away_score", "")
            if home_score != "" and away_score != "":
                try:
                    match.home_score = int(home_score)
                    match.away_score = int(away_score)
                except (ValueError, TypeError):
                    pass

            is_finished_raw = str(row.get("is_finished", "")).strip().upper()
            if is_finished_raw in ("TRUE", "1", "YES"):
                match.is_finished = True
            elif is_finished_raw in ("FALSE", "0", "NO"):
                match.is_finished = False

            adv_code = row.get("advanced_team_code", "").strip()
            if adv_code:
                t = get_team(adv_code)
                if t:
                    match.advanced_team_id = t.id

            match.result_note = row.get("result_note", match.result_note)

        except Exception as e:
            errors.append(f"Kamp {row.get('match_number', '?')}: {e}")

    if errors:
        db.session.rollback()
        return {"success": False, "errors": errors}

    db.session.commit()
    db.session.add(AdminAuditLog(
        action="excel_csv_import",
        details=f"Nye: {new_count}, oppdatert: {updated_count}"
    ))
    db.session.commit()
    recalculate_all_scores()
    return {"success": True, "new": new_count, "updated": updated_count, "errors": []}


def _parse_import_file(content, file_type):
    rows = []
    errors = []
    try:
        if file_type == "xlsx":
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [str(cell.value or "").strip() for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if all(v is None for v in row):
                    continue
                rows.append({headers[i]: (str(v) if v is not None else "") for i, v in enumerate(row) if i < len(headers)})
        elif file_type == "csv":
            text = content.decode("utf-8-sig") if isinstance(content, bytes) else content
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
    except Exception as e:
        errors.append(f"Kunne ikke lese fil: {e}")
    return rows, errors


def _validate_row(row):
    mn = row.get("match_number", "")
    if not mn:
        return "match_number mangler"
    try:
        int(mn)
    except (ValueError, TypeError):
        return f"match_number '{mn}' er ikke et tall"
    phase = row.get("phase", "").strip().lower()
    if phase not in ("group", "knockout", ""):
        return f"phase '{phase}' er ugyldig"
    return None


def export_users_csv():
    """Export all users and their scores to CSV."""
    from scoring import get_user_total_points
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["id", "name", "email", "created_at", "last_seen_at", "total_points", "is_active"])
    for user in User.query.all():
        writer.writerow([
            user.id, user.name, user.email,
            user.created_at, user.last_seen_at,
            get_user_total_points(user.id),
            user.is_active
        ])
    return buf.getvalue().encode("utf-8")
